'''
Dateiname: ebay Scraper

Beschreibung: Durchsucht eBay nach den letzten Verkäufen und speichert die Ausgaben in einzelnen Excel-Dateien im (ggf.)
neu zu erstellenden Ordner "Preisverläufe". Die Ergebnisse werden überdies im Terminal ausgegeben. Es wird
berücksichtigt, dass Daten die nicht doppelt erhoben werden.
Das Datum ist im Schema JJJJ.MM.TT aufgebaut. Die Tabellen werden mit jedem Durchlauf nach Datum aufsteigend sortiert
und der Durchschnittspreis sowie der Höchstverkaufspreis wird gespeichert.
Die URLs können entweder in Zeile 249 manuell kopiert werden, oder es wird nur der Name angepasst. Die Sucheinstellungen
betreffen Verkäufe in der Europäischen Union. Mehrere URLs müssen in der Liste durch ein Komma getrennt werden.
Die Excel-Listen können abschließend um Fehlerfassungen bereinigt werden. Dazu kann die URL direkt aus der Excel-Datei
geöffnet werden.

Ebay speichert die Verkaufsdaten der letzten 3 Monate. Das Script sollte daher bestenfalls alle 3 Monate ausgeführt
werden, damit keine Fehlerfassung nach der manuellen Löschung erneut erfasst werden.

Autor: d1ggdugg
Unterstützung bei der Entwicklung durch ChatGPT (OpenAI)

Version: 1.0
Stand: 22.06.2025
'''

import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os

def datei_scraping(url):

	# Kartenbezeichnung (Set und Nummer oder Name) und PSA-Grade als Titel angeben
	treffer = re.search(r'_nkw=([^&]+)&LH_', url)
	if treffer:
		kartenname = url.split("_nkw=")[1].split("&LH_")[0]  # sm11+011+psa+9
		kartenname = kartenname.replace("+", " ").replace("psa", "PSA") # sm11 011 PSA 9
		print(f"\n{kartenname}")

	# URL als weiteren Titel ausgeben
	print(f"{url}")

	# Prüfung der http-Anfrage und Definierung von response
	response = requests.get(url)
	response.raise_for_status()

	# Soup-Standard
	soup = BeautifulSoup(response.text, 'html.parser')

	# Prüfung, ob überhaupt passende Artikel gefunden wurden
	# Nur wenn es keinen Artikel gibt, gibt es srp-save-null-search
	keine_suchergebnisse = soup.find("div", class_="srp-save-null-search")
	if keine_suchergebnisse:
		print("Keine Suchergebnisse gefunden.")

	else:
		# Finden von unsortierten Listen (ul), das sind die gefundenen Suchergebnisse
		ul_element = soup.find("ul", class_="srp-results srp-list clearfix")

		for li in ul_element.find_all("li", recursive=False):

			# Sobald in den Suchergebnissen "Ergebnisse für weniger Suchbegriffe" auftaucht, wird der Vorgang beendet
			classes = li.get("class", [])
			if "srp-river-answer" in classes and "srp-river-answer--REWRITE_START" in classes:
				break

			# Extrahiert die Rohdaten bzgl. Verkaufsdatum,  Verkaufspreis und Versandkosten
			datum_element = li.find(class_="s-item__caption--signal POSITIVE")
			preis_element = li.find(class_="s-item__price")
			versand_element = li.find(class_="s-item__shipping s-item__logisticsCost") or li.find(
				class_="s-item__dynamic s-item__paidDeliveryInfo")

			# Formatierung des Datums zu YYYY.MM.TT (2025.01.01)
			date_str = datum_element.get_text(strip=True) if datum_element else "01. Jan 1900"
			date_str = date_str.replace("Verkauft", "")
			date_str = date_str.strip()

			# Ersetze den deutschen Monatsnamen durch die englische Abkürzung
			monate = {
				"Jan": "Jan", "Feb": "Feb", "Mrz": "Mar", "Apr": "Apr", "Mai": "May", "Jun": "Jun",
				"Jul": "Jul", "Aug": "Aug", "Sep": "Sep", "Okt": "Oct", "Nov": "Nov", "Dez": "Dec"
			}

			for de, en in monate.items():
				if de in date_str:
					date_str = date_str.replace(de, en)
					break  # Nur den ersten Treffer ersetzen

			# Umwandlung in ein "datetime"-Objekt
			date_obj = datetime.strptime(date_str, "%d. %b %Y")

			# Umformatierung zu JJJJ.MM.TT
			formatiertes_datum = date_obj.strftime("%Y.%m.%d")

			# Formatierung des Verkaufspreises
			preis = preis_element.get_text(strip=True) if preis_element else "Kein Preis gefunden"
			preis = preis.replace("EUR", "")
			preis = preis.lstrip()

			try:
				preis_float = float(preis.replace(",", "."))
			except ValueError:
				preis_float = 0.0

			# Formatierung des Versandpreises
			if versand_element:
				versand = versand_element.get_text(strip=True)

				versand = versand.replace("\xa0", " ")
				versand = re.sub(r"[+·]", "", versand)
				versand = re.sub(r"EUR|Lieferung|2-3 Tage Lieferung", "", versand)
				versand = re.sub(r"Gratis", "0,00", versand)
				versand = re.sub(r"·", "", versand)

				versand = versand.strip()
				versand = versand.lstrip()
				versand = versand.rstrip()

			else:
				alt_versand_element = li.find(class_="s-item__dynamic s-item__paidDeliveryInfo")
				versand = alt_versand_element.get_text(strip=True) if alt_versand_element else "Keine Versandkosten gefunden"

			try:
				versand_float = float(versand.replace(",", "."))
			except ValueError:
				versand_float = 0.0

			# Speicherung der Daten in die Liste "verkaufsdaten"
			if formatiertes_datum == "1900.01.01":
				pass
			else:
				verkaufsdaten = []

				# Eintrag zur Liste hinzufügen
				verkaufsdaten.append({
					"datum": formatiertes_datum,
					"preis": preis_float,
					"versand": versand_float,
					"gesamt": preis_float + versand_float,
				})

				# Zielordner erstellen
				ordnername = "Preisverläufe"
				os.makedirs(ordnername, exist_ok=True)

				# Definition des Datennamens (= Kartenname)
				dateiname = os.path.join(ordnername, f"{kartenname}.xlsx")

				# Definition für das Datum in der Excel-Datei
				def parse_date(datum):
					if isinstance(datum, datetime):
						return datum
					else:
						return datetime.strptime(datum, "%Y.%m.%d")

				# Datei öffnen oder neu erstellen
				neu_erstellt = not os.path.exists(dateiname)
				if not neu_erstellt:
					wb = load_workbook(dateiname)
					ws = wb.active
				else:
					wb = Workbook()
					ws = wb.active
					ws["A1"].value = url
					ws["A1"].hyperlink = url
					ws["A1"].style = "Hyperlink"
					ws.append([])  # Leere Zeile 2
					ws.append(["Datum", "Preis", "Versand", "Gesamtpreis"])

				# Bestehende Daten ab Zeile 4 einlesen
				def normalisiere_eintrag(row):
					datum_str = row[0].strftime("%Y.%m.%d") if isinstance(row[0], datetime) else str(row[0])
					return (
						datum_str,
						round(float(row[1]), 2) if row[1] is not None else 0.0,
						round(float(row[2]), 2) if row[2] is not None else 0.0,
						round(float(row[3]), 2) if row[3] is not None else 0.0
					)

				bestehende_daten = set(
					normalisiere_eintrag(row)
					for row in ws.iter_rows(min_row=4, max_col=4, values_only=True)
				)

				# Neue Daten anhängen, wenn nicht vorhanden
				for eintrag in verkaufsdaten:
					eintrags_tuple = (
						eintrag["datum"],
						round(eintrag["preis"], 2),
						round(eintrag["versand"], 2),
						round(eintrag["gesamt"], 2)
					)
					if eintrags_tuple not in bestehende_daten:
						ws.append(list(eintrags_tuple))

				# Daten lesen und nach Datum sortieren
				daten_zeilen = list(ws.iter_rows(min_row=4, max_col=4, values_only=True))
				daten_zeilen.sort(key=lambda x: parse_date(x[0]))

				# Alte Daten löschen (ab Zeile 4)
				for _ in range(ws.max_row - 3):
					ws.delete_rows(4)

				# Sortierte Daten wieder einfügen
				for row in daten_zeilen:
					datum_dt = parse_date(row[0])
					ws.append([datum_dt, row[1], row[2], row[3]])

				# Formatierung Datum & Euro
				for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=4):
					row[0].number_format = 'yyyy.mm.dd'
					for cell in row[1:4]:
						cell.number_format = u'#,##0.00 €'

				# Durchschnitt und Höchstpreis berechnen
				gesamtpreise = [row[3] for row in daten_zeilen if isinstance(row[3], (int, float))]
				durchschnitt = round(sum(gesamtpreise) / len(gesamtpreise), 2) if gesamtpreise else 0
				maximum = round(max(gesamtpreise), 2) if gesamtpreise else 0

				# Ergebnisse in E3/F3 und E4/F4
				ws["E3"] = "Durchschnittspreis"
				ws["E4"] = "=AVERAGE(D4:D1000)"
				ws["F3"] = "Höchstpreis"
				ws["F4"] = "=MAX(D4:D1000)"
				ws["E4"].number_format = u'#,##0.00 €'
				ws["F4"].number_format = u'#,##0.00 €'

				# Überschriften fett formatieren
				for cell in ws[3]:
					cell.font = Font(bold=True)

				# Spaltenbreite fest auf 19
				for col in ws.columns:
					col_letter = get_column_letter(col[0].column)
					ws.column_dimensions[col_letter].width = 19

				# Zellen linksbündig ausrichten (ab Zeile 4)
				for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
					for cell in row:
						cell.alignment = Alignment(horizontal='left')

				# Speichern
				wb.save(dateiname)
				print(f"{formatiertes_datum};{preis};{versand}")

urls = ["https://www.ebay.de/sch/i.html?_nkw=sm11+011+psa+10&LH_Sold=1&LH_Complete=1&rt=nc&LH_PrefLoc=3",
		"https://www.ebay.de/sch/i.html?_nkw=sm11+011+psa+10&LH_Sold=1&LH_Complete=1&rt=nc&LH_PrefLoc=3",
		"https://www.ebay.de/sch/i.html?_nkw=sm11+011+psa+10&LH_Sold=1&LH_Complete=1&rt=nc&LH_PrefLoc=3",
	]

for url in urls:
	datei_scraping(url)
